import calendar
from datetime import date
from django.utils import translation
from django.http import HttpResponse
from django.urls import reverse, reverse_lazy
from django.shortcuts import redirect, render, get_object_or_404
from django.core.exceptions import ValidationError
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.translation import gettext as _
from django.utils.decorators import method_decorator
from django.db.models import Q
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView, TemplateView, DetailView
)
from Quran.models import (
    School, Student, Teacher, Course, ClassGroup, Attendance, Invoice, StudentPaymentStatus
)
from Quran.forms import (
    StudentForm, TeacherForm, CourseForm, ClassGroupForm
)
from Quran.utils import (
    get_user_school, get_present_for_student, get_attendance_summary, get_missing_months_for_student, get_group_summary, get_payment_summary,
)
from Quran.services import (
    handle_academic_year
)

# Excle Importing
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from django.contrib.staticfiles import finders
from openpyxl.styles import Alignment, Font, PatternFill
from django.conf import settings

# Pdf Importing
from weasyprint import HTML
from django.template.loader import render_to_string


# --------------------- Home ---------------------
@method_decorator(staff_member_required, name='dispatch')
class HomeView(TemplateView):
    template_name = 'Quran/home.html'

    # Everytime will check date to upgrade academic year
    today = date.today()
    if today.month == 9 and today.day == 1:
        handle_academic_year()


# --------------------- Base Class ---------------------
class BaseListView(ListView):
    def get_queryset(self):
        queryset = super().get_queryset()
        user_schools = get_user_school(self.request)
        if user_schools.exists():
            queryset = queryset.filter(school__in=user_schools)
        return queryset.order_by('-updated_at')


class BaseCreateView(CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'Quran/student/student_form.html'
    success_url = reverse_lazy('student_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser:
            form.instance.school = get_user_school(self.request).first()
        return super().form_valid(form)


class BaseUpdateView(UpdateView):
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        if not self.request.user.is_superuser:
            form.instance.school = get_user_school(self.request).first()
        return super().form_valid(form)


# --------------------- Student ---------------------
@method_decorator(staff_member_required, name='dispatch')
class StudentListView(BaseListView):
    model = Student
    template_name = 'Quran/student/student_list.html'
    context_object_name = 'students'
    paginate_by = 12

    def get_queryset(self):
        queryset = super().get_queryset()

        # Get search query from the URL
        search_query = self.request.GET.get('q', '')

        # Get gender and group filters
        gender_filter = self.request.GET.get('gender', '')
        group_filter = self.request.GET.get('group', '')

        # Apply search filter by name or code
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | Q(code__icontains=search_query)
            )

        # Apply gender filter
        if gender_filter:
            queryset = queryset.filter(gender=gender_filter)

        # Apply group filter
        if group_filter:
            queryset = queryset.filter(group=group_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Pass group options 
        user_school = get_user_school(self.request)
        context['groups'] = ClassGroup.objects.filter(school__in=user_school)

        # Pass the current search query and gender and group filter options to the context
        context['search_query'] = self.request.GET.get('q', '')
        context['gender_filter'] = self.request.GET.get('gender', '')
        context['group_filter'] = self.request.GET.get('group', '')

        return context


@method_decorator(staff_member_required, name='dispatch')
class StudentCreateView(BaseCreateView):
    model = Student
    form_class = StudentForm
    template_name = 'Quran/student/student_form.html'
    success_url = reverse_lazy('student_list')


@method_decorator(staff_member_required, name='dispatch')
class StudentUpdateView(BaseUpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'Quran/student/student_form.html'
    success_url = reverse_lazy('student_list')


@method_decorator(staff_member_required, name='dispatch')
class StudentDeleteView(DeleteView):
    model = Student
    template_name = 'Quran/student/student_confirm_delete.html'
    success_url = reverse_lazy('student_list')


# --------------------- Teacher ---------------------
@method_decorator(staff_member_required, name='dispatch')
class TeacherListView(BaseListView):
    model = Teacher
    template_name = 'Quran/teacher/teacher_list.html'
    context_object_name = 'teachers'
    paginate_by = 12

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | Q(id_number__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context

@method_decorator(staff_member_required, name='dispatch')
class TeacherCreateView(BaseCreateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'Quran/teacher/teacher_form.html'
    success_url = reverse_lazy('teacher_list')


@method_decorator(staff_member_required, name='dispatch')
class TeacherUpdateView(BaseUpdateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'Quran/teacher/teacher_form.html'
    success_url = reverse_lazy('teacher_list')


@method_decorator(staff_member_required, name='dispatch')
class TeacherDeleteView(DeleteView):
    model = Teacher
    template_name = 'Quran/teacher/teacher_confirm_delete.html'
    success_url = reverse_lazy('teacher_list')


# --------------------- Course ---------------------
from django.contrib.auth.decorators import permission_required

@method_decorator(permission_required('Quran.view_course', raise_exception=True), name='dispatch')
class CourseListView(BaseListView):
    model = Course
    template_name = 'Quran/course/course_list.html'
    context_object_name = 'courses'
    paginate_by = 12

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                name__icontains=search_query
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context

@method_decorator(permission_required('Quran.add_course', raise_exception=True), name='dispatch')
class CourseCreateView(BaseCreateView):
    model = Course
    form_class = CourseForm
    template_name = 'Quran/course/course_form.html'
    success_url = reverse_lazy('course_list')


@method_decorator(permission_required('Quran.change_course', raise_exception=True), name='dispatch')
class CourseUpdateView(BaseUpdateView):
    model = Course
    form_class = CourseForm
    template_name = 'Quran/course/course_form.html'
    success_url = reverse_lazy('course_list')


@method_decorator(permission_required('Quran.delete_course', raise_exception=True), name='dispatch')
class CourseDeleteView(DeleteView):
    model = Course
    template_name = 'Quran/course/course_confirm_delete.html'
    success_url = reverse_lazy('course_list')


# --------------------- ClassGroup ---------------------
@method_decorator(staff_member_required, name='dispatch')
class ClassGroupListView(BaseListView):
    model = ClassGroup
    template_name = 'Quran/class_group/class_group_list.html'
    context_object_name = 'class_groups'
    paginate_by = 12

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | Q(code__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context

@method_decorator(staff_member_required, name='dispatch')
class ClassGroupCreateView(BaseCreateView):
    model = ClassGroup
    form_class = ClassGroupForm
    template_name = 'Quran/class_group/class_group_form.html'
    success_url = reverse_lazy('class_group_list')


@method_decorator(staff_member_required, name='dispatch')
class ClassGroupUpdateView(BaseUpdateView):
    model = ClassGroup
    form_class = ClassGroupForm
    template_name = 'Quran/class_group/class_group_form.html'
    success_url = reverse_lazy('class_group_list')


@method_decorator(staff_member_required, name='dispatch')
class ClassGroupDeleteView(DeleteView):
    model = ClassGroup
    template_name = 'Quran/class_group/class_group_confirm_delete.html'
    success_url = reverse_lazy('class_group_list')


# --------------------- Attendance ---------------------
@method_decorator(staff_member_required, name='dispatch')
class AttendanceView(TemplateView):
    template_name = 'Quran/attendance/attendance.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        # Pass group options 
        user_school = get_user_school(self.request)
        groups = ClassGroup.objects.filter(school__in=user_school)

        context.update({
            'today': today,
            'groups': groups
        })

        # Add post-processing context if available
        if hasattr(self, "post_context"):
            context.update(self.post_context)

        return context

    def get(self, request, *args, **kwargs):
        selected_date = self.request.GET.get('selected_date')
        selected_group = self.request.GET.get('selected_group')

        # Fetch group students
        group_students = self.get_group_students(self.request, selected_date, selected_group)

        self.post_context = {
            'selected_date': selected_date,
            'selected_group': selected_group,
            'group_students': group_students,
        }

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        selected_date = request.POST.get('selected_date')
        selected_group = request.POST.get('selected_group')
        student_ids = request.POST.getlist('student_ids')
        present_student_ids = request.POST.getlist('present')

        # Update attendance records for the selected date and group
        for student_id in student_ids:
            present = str(student_id) in present_student_ids
            student = get_object_or_404(Student, id=int(student_id))

            Attendance.objects.update_or_create(
                school=student.school,
                student=student,
                date=selected_date,
                defaults={'present': present}
            )

        # Fetch group students
        group_students = self.get_group_students(self.request, selected_date, selected_group)

        self.post_context = {
            'selected_date': selected_date,
            'selected_group': selected_group,
            'group_students': group_students,
        }

        return super().get(request, *args, **kwargs)

    def get_group_students(self, request, date, group):
        user_school = get_user_school(request)

        group_students = []
        if date and group:
            students = Student.objects.filter(is_active=True, school__in=user_school, group=int(group))
            for student in students:
                missing = get_missing_months_for_student(student)
                present = get_present_for_student(student.id, date)
                group_students.append({
                    "student_id": student.id,
                    "student_code": student.code,
                    "student_name": student.name,
                    "status": missing,
                    "present": present,
                })

        return group_students


@method_decorator(staff_member_required, name='dispatch')
class MonthlyAttendanceView(TemplateView):
    template_name = 'Quran/attendance/monthly_attendance.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        
        context.update({
            "schools": School.objects.all(),
            "selected_month": int(today.month),
            "selected_year": int(today.year),
            "months": range(1, 13),
            "years": range(today.year, today.year + 10),
        })
        
        # Add post-processing context if available
        if hasattr(self, "post_context"):
            context.update(self.post_context)
        
        return context

    def get(self, request, *args, **kwargs):
        # Get filter parameters
        school_id = request.GET.get('school')
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        # Process and get attendance data
        attendance_data = self.get_attendance_data(school_id, month, year)
        
        # Store in post_context for template
        self.post_context = attendance_data
        
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Read action: filter or pdf
        action = request.POST.get("action")
        
        # Get filter parameters
        school_id = request.POST.get('school')
        month = request.POST.get('month')
        year = request.POST.get('year')
        
        # Process and get attendance data
        attendance_data = self.get_attendance_data(school_id, month, year)
        
        # Export to PDF if requested
        if action == "pdf":
            return self.export_pdf(request, attendance_data)
        
        # Store in post_context for template rendering
        self.post_context = attendance_data
        
        # Otherwise, render the filtered report page
        return super().get(request, *args, **kwargs)

    def get_attendance_data(self, school_id=None, month=None, year=None):
        """Process and return attendance data based on filters."""
        today = date.today()
        data = None
        days = []

        # Get school
        if not self.request.user.is_superuser and not school_id:
            school_id = get_user_school(self.request).first().id

        # Convert values
        school_id = int(school_id) if school_id else None
        month = int(month) if month else today.month
        year = int(year) if year else today.year

        # Get attendance data if a school is selected
        if school_id:
            data = get_attendance_summary(
                school_id=school_id,
                month=month,
                year=year
            )
            
            # Calculate days in month for calendar display
            last_day = calendar.monthrange(year, month)[1]
            days = range(1, last_day + 1)
        
        context = {
            "selected_school": school_id,
            "selected_month": month,
            "selected_year": year,
            "days": days,
            "data": data,
        }

        return context

    def export_pdf(self, request, attendance_data):
        """Export attendance data to PDF."""
        
        # Render HTML template with data
        html_string = render_to_string('Quran/attendance/monthly_attendance_pdf.html', attendance_data)
        
        # Create PDF from HTML
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        
        # Generate PDF
        pdf_file = html.write_pdf()
        
        # Create HTTP response with PDF
        response = HttpResponse(pdf_file, content_type='application/pdf')
        
        # Set filename
        filename = f"attendance_report_{attendance_data['selected_month']}_{attendance_data['selected_year']}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


# --------------------- Invoices ---------------------
@method_decorator(staff_member_required, name='dispatch')
class InvoiceCreateView(TemplateView):
    template_name = 'Quran/invoice/invoice.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        current_month = today.month
        current_year = today.year
        
        context.update({
            "months": range(1, 13),
            "years": range(current_year, current_year + 10),
            "current_month": current_month,
            "current_year": current_year,
        })
        
        # Add post-processing context if available
        if hasattr(self, "post_context"):
            context.update(self.post_context)
            
        return context

    def get(self, request, *args, **kwargs):
        selected_code = request.GET.get("selected_code", '')
        selected_name = request.GET.get("selected_name", '')
        
        self.post_context = self.process_student_data(code=selected_code, name=selected_name)
        
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        code = request.POST.get('student_code')
        year = request.POST.get('year')
        month = request.POST.get('month')
        
        # First, get student data context
        self.post_context = self.process_student_data(code=code)
        
        # If we have student data (no errors), try to create invoice
        if self.post_context.get('student_data'):
            student_id = self.post_context['student_data']['id']
            student = get_object_or_404(Student, id=student_id)
            
            # Check for duplicate invoice
            if Invoice.objects.filter(student=student, month=month, year=year).exists():
                self.post_context["error_msg"] = _(f"An invoice already exists for this student for this month.")

            else:
                # Create invoice
                try:
                    invoice = Invoice.objects.create(
                        school=student.school,
                        student=student,
                        month=month,
                        year=year,
                        amount=Invoice.calculate_expected_amount(student),
                    )
                    return redirect("print_invoice", invoice_id=invoice.id)
                except ValidationError as e:
                    self.post_context["error_msg"] = _("Invoice was not saved.")
        
        return super().get(request, *args, **kwargs)

    def process_student_data(self, code=None, name=None):
        """Process and return context for a student. Handles validation and data building."""
        context = {}
        
        if code:
            context['selected_code'] = code
        if name:
            context['selected_name'] = name
            
        # Only process if we have search criteria
        if not (code or name):
            return context
            
        # Get student
        student = None
        if code:
            student = Student.objects.filter(code=code).first()
        elif name:
            student = Student.objects.filter(name__icontains=name).first()

        # Get student data
        if not student:
            context["error_msg"] = _("Student not found.")
        elif student.discount_type == 'full':
            context["error_msg"] = _("Student is exempt.")
        else:
            context["student_data"] = self.build_student_context(student)
            
        return context

    def build_student_context(self, student):
        """Return context dictionary for a student."""
        expected_price = Invoice.calculate_expected_amount(student)
        missing = get_missing_months_for_student(student)
        return {
            "id": student.id,
            "code": student.code,
            "name": student.name,
            "group": student.group.name,
            "course": student.group.course.name,
            "price": int(expected_price),
            "status": missing,
        }


@method_decorator(staff_member_required, name='dispatch')
class InvoicePrintView(TemplateView):
    template_name = 'Quran/invoice/print_invoice.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        invoice_id = self.kwargs.get("invoice_id")
        invoice = get_object_or_404(Invoice, pk=invoice_id)

        # Get missing months
        missing = get_missing_months_for_student(invoice.student)

        # Update context with all required data for template
        context.update({
            "id": invoice.id,
            "date": invoice.date,
            "student": invoice.student.name,
            "course": invoice.student.group.course.name,
            "month": invoice.month,
            "year": invoice.year,
            "amount": int(invoice.amount),
            "status": missing,
        })

        context["invoice"] = invoice
        return context


# --------------------- Student Payment Status ---------------------
@method_decorator(staff_member_required, name='dispatch')
class PaymentStatusListView(TemplateView):
    template_name = 'Quran/student_payment_status/student_payment_status_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        # Add schools and current date to context
        context.update({
            "today": str(today),
            "from_date": str(today),
            "to_date": str(today),
            "schools": School.objects.all(),
        })

        # Add post-processing context if available
        if hasattr(self, "post_context"):
            context.update(self.post_context)

        return context

    # POST generate or exporting Excel
    def post(self, request, *args, **kwargs):
        # Extract and process the form values
        action = request.POST.get("action")
        school_id = self.request.POST.get("school")
        from_date = self.request.POST.get("from_date")
        to_date = self.request.POST.get("to_date")

        # Fetch the data based on the selected filters
        data, totals = self.get_payment_data(school_id, from_date, to_date)

        # Store the filtered data and totals for use in the template
        self.post_context = {
            "selected_school": school_id,
            "from_date": from_date,
            "to_date": to_date,
            "data": data,
            "totals": totals,
        }

        # Export to Excel if requested
        if action == "excel" and data:
            return self.export_excel(data)

        # Otherwise, render the filtered report page
        return self.get(request, *args, **kwargs)

    def get_payment_data(self, school_id=None, from_date=None, to_date=None):
        data = None
        totals = {}

        # Convert values to integers
        if not self.request.user.is_superuser and not school_id:
            school_id = get_user_school(self.request).first()
        school_id = int(school_id)

        # Fetch the payment summary for the given filters
        if school_id:
            data = get_payment_summary(school_id=school_id, from_date=from_date, to_date=to_date)

            # Calculate totals for numeric columns
            if data:
                totals = {"total_amount": sum(d.get("amount", 0) for d in data)}

        return data, totals

    # Export Excel
    def export_excel(self, data):
        """
        Generates an Excel file from the payment status data and returns it as a response.
        Includes logo, merged title, striped header, full column width, and row height.
        """
        wb = Workbook()
        ws = wb.active

        # Determine language
        lang = translation.get_language()
        
        # Set headers based on language
        if lang == "ar":
            title_text = "حالة الدفع"
            headers = [
                "التاريخ", "رمز الطالب", "اسم الطالب", "الشهر", "السنة", "المبلغ"
            ]
            ws.sheet_view.rightToLeft = True
        else:
            title_text = _("Payment Status")
            headers = [
                _("Date"), _("Student Code"), _("Student Name"), _("Month"), _("Year"), _("Amount"),
            ]

        # Add logo in first row
        # Get the absolute path to the static file
        logo_path = finders.find("logo.png")
        if logo_path:
            img = Image(logo_path)
            img.height = 60
            img.width = 120
            ws.add_image(img, "A1")

        # Merge cells for title
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=2)
        title_cell.value = title_text
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(size=14, bold=True)
        ws.row_dimensions[1].height = 48

        # Header row
        header_row_index = 2
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_index, column=col_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row_index].height = 25

        # Data rows with stripe effect
        for i, row in enumerate(data, start=header_row_index + 1):
            fill = PatternFill(start_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                                end_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                                fill_type="solid")
            
            row_values = [
                row.get("date"),
                row.get("student_code"),
                row.get("student_name"),
                row.get("month"),
                row.get("year"),
                row.get("amount"),
            ]
            
            for col_index, value in enumerate(row_values, start=1):
                cell = ws.cell(row=i, column=col_index, value=value)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[i].height = 30

        # Totals row (if applicable)
        if data:
            totals_row_index = ws.max_row + 1
            totals_row_values = [
                _("Totals"), "", "", "", "",
                sum(d.get("amount", 0) for d in data),
            ]
            for col_index, value in enumerate(totals_row_values, start=1):
                cell = ws.cell(row=totals_row_index, column=col_index, value=value)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws.row_dimensions[totals_row_index].height = 25

        # Adjust column widths
        for i, col in enumerate(ws.iter_cols(min_row=header_row_index, max_row=ws.max_row), start=1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except AttributeError:
                    continue  # Skip merged cells
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save and return response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        export_filename = _("payment_status.xlsx")
        response["Content-Disposition"] = f'attachment; filename="{export_filename}"'
        wb.save(response)
        return response


# --------------------- Reports ---------------------
@method_decorator(staff_member_required, name='dispatch')
class SummaryReportView(TemplateView):
    template_name = 'Quran/reports/summary_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        
        context.update({
            "schools": School.objects.all(),
            "selected_month": int(today.month),
            "selected_year": int(today.year),
            "months": range(1, 13),
            "years": range(today.year, today.year + 10),
        })
        
        # Add post-processing context if available
        if hasattr(self, "post_context"):
            context.update(self.post_context)
            
        return context
    
    def get(self, request, *args, **kwargs):
        # Get filter parameters
        school_id = request.GET.get('school')
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        # Process and get report data
        report_data = self.get_report_data(school_id, month, year)
        
        # Store in post_context for template
        self.post_context = report_data
        
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Read action: filter or excel
        action = request.POST.get("action")
        
        # Get filter parameters
        school_id = request.POST.get('school')
        month = request.POST.get('month')
        year = request.POST.get('year')
        
        # Process and get report data
        report_data = self.get_report_data(school_id, month, year)
        
        # Export to Excel if requested
        if action == "excel" and report_data.get("data"):
            return self.export_excel(report_data["data"])
        
        # Store in post_context for template rendering
        self.post_context = report_data
        
        # Otherwise, render the filtered report page
        return super().get(request, *args, **kwargs)

    def get_report_data(self, school_id=None, month=None, year=None):
        """Process and return report data based on filters."""
        today = date.today()
        data = None
        totals = {}

        # Convert and validate parameters
        school_id = int(school_id) if school_id else None
        month = int(month) if month else today.month
        year = int(year) if year else today.year

        # Get report data if a school is selected
        if school_id:
            data = get_group_summary(
                school_id=school_id,
                month=month,
                year=year
            )

            # Calculate totals for numeric columns
            if data:
                totals = {
                    "total_students": sum(d.get("total_students", 0) for d in data),
                    "students_paid_current": sum(d.get("students_paid_current", 0) for d in data),
                    "students_discount_current": sum(d.get("students_discount_current", 0) for d in data),
                    "students_full_discount": sum(d.get("students_full_discount", 0) for d in data),
                    "students_not_paid": sum(d.get("students_not_paid", 0) for d in data),
                    "students_paid_previous": sum(d.get("students_paid_previous", 0) for d in data),
                    "students_discount_previous": sum(d.get("students_discount_previous", 0) for d in data),
                    "final_total": sum(d.get("final_total", 0) for d in data),
                }

        # Update context with all required data for template
        context = {
            "selected_school": school_id,
            "selected_month": month,
            "selected_year": year,
            "data": data,
            "totals": totals,
        }
        return context

    # Export Excel
    def export_excel(self, data):
        """
        Generates an Excel file from the summary report data and returns it as a response.
        Includes logo, merged title, striped header, full column width, and row height.
        """
        wb = Workbook()
        ws = wb.active

        # Determine language
        lang = translation.get_language()
        
        # Set headers based on language
        if lang == "ar":
            title_text = "تقرير ملخص"
            headers = [
                "المعلم", "المجموعة", "البداية", "النهاية", "إجمالي الطلاب",
                "المدفوع هذا الشهر", "الخصم هذا الشهر", "إعفاء كامل", "غير مدفوع",
                "المدفوع الأشهر السابقة", "الخصم الأشهر السابقة", "الإجمالي"
            ]
            ws.sheet_view.rightToLeft = True
        else:
            title_text = _("Summary Report")
            headers = [
                _("Teacher"), _("Group"), _("Start"), _("End"), _("Total Students"),
                _("Paid Current"), _("Discount Current"), _("Full Discount"), _("Not Paid"),
                _("Paid Previous"), _("Discount Previous"), _("Total Amount")
            ]

        # Add logo in first row
        # Get the absolute path to the static file
        logo_path = finders.find("logo.png")
        if logo_path:
            img = Image(logo_path)
            img.height = 60
            img.width = 120
            ws.add_image(img, "A1")

        # Merge cells for title
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=2)
        title_cell.value = title_text
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(size=14, bold=True)
        ws.row_dimensions[1].height = 48

        # Header row
        header_row_index = 2
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_index, column=col_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row_index].height = 25

        # Data rows with stripe effect
        for i, row in enumerate(data, start=header_row_index + 1):
            fill = PatternFill(start_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                            end_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                            fill_type="solid")
            row_values = [
                row.get("teacher__name"),
                row.get("name"),
                row.get("start_time"),
                row.get("end_time"),
                row.get("total_students"),
                row.get("students_paid_current"),
                row.get("students_discount_current"),
                row.get("students_full_discount"),
                row.get("students_not_paid"),
                row.get("students_paid_previous"),
                row.get("students_discount_previous"),
                row.get("final_total"),
            ]
            for col_index, value in enumerate(row_values, start=1):
                cell = ws.cell(row=i, column=col_index, value=value)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[i].height = 30

        # Totals row
        totals_row_index = ws.max_row + 1
        totals_row_values = [
            _("Totals"), "", "", "",
            sum(d.get("total_students", 0) for d in data),
            sum(d.get("students_paid_current", 0) for d in data),
            sum(d.get("students_discount_current", 0) for d in data),
            sum(d.get("students_full_discount", 0) for d in data),
            sum(d.get("students_not_paid", 0) for d in data),
            sum(d.get("students_paid_previous", 0) for d in data),
            sum(d.get("students_discount_previous", 0) for d in data),
            sum(d.get("final_total", 0) for d in data),
        ]
        for col_index, value in enumerate(totals_row_values, start=1):
            cell = ws.cell(row=totals_row_index, column=col_index, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws.row_dimensions[totals_row_index].height = 25

        # Adjust column widths
        for i, col in enumerate(ws.iter_cols(min_row=header_row_index, max_row=ws.max_row), start=1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except AttributeError:
                    continue  # Skip merged cells
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save and return response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        export_filename = _("summary_report.xlsx")
        response["Content-Disposition"] = f'attachment; filename="{export_filename}"'
        wb.save(response)
        return response

